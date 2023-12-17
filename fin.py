#!/usr/bin/env python
# coding: utf-8

# In[1]:


# get_ipython().system('pip install requests')
# get_ipython().system('pip install openpyxl')


# In[2]:


import requests
import json
import logging
import openpyxl
import time
from openpyxl.styles import *


# In[3]:


logger = logging.getLogger()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
 
 
# Configure stream handler for the cells
chandler = logging.StreamHandler()
chandler.setLevel(logging.WARN)
chandler.setFormatter(formatter)
 
# Add both handlers
# logger.addHandler(fhandler)
logger.addHandler(chandler)
logger.setLevel(logging.WARN)


# In[7]:


stock_code=  input('请输入股票代码，如 SH600000\n')
token ='需要填入雪球的真实xq_a_token'
count=  input('输入报表数量，默认10\n')
if len(count) == 0:
    count=10
print("报表数量： "+str(count))
def fetch(url, host="stock.xueqiu.com"):
    HEADERS = {'Host': host,
               'Accept': 'application/json',
               'Cookie': 'xq_a_token='+token,
               'User-Agent': 'Xueqiu iPhone 11.8',
               'Accept-Language': 'zh-Hans-CN;q=1, ja-JP;q=0.9',
               'Accept-Encoding': 'br, gzip, deflate',
               'Connection': 'keep-alive'}

    response = requests.get(url, headers=HEADERS)

    if response.status_code != 200:
        raise Exception(response.content)
    return json.loads(response.content)


def fetch_balance(code):
    timestamp = round(time.time()*1000)
    url = 'https://stock.xueqiu.com/v5/stock/finance/cn/balance.json?symbol={}&type=all&is_detail=true&count={}&timestamp={}'.format(code,count,timestamp)
    logger.info("Request ur is "+url)
    return fetch(url)

def fetch_income(code):
    timestamp = round(time.time()*1000)
    url = 'https://stock.xueqiu.com/v5/stock/finance/cn/income.json?symbol={}&type=all&is_detail=true&count={}&timestamp={}'.format(code,count,timestamp)
    logger.info("Request ur is "+url)
    return fetch(url)
def fetch_indicator(code):
    timestamp = round(time.time()*1000)
    url = 'https://stock.xueqiu.com/v5/stock/finance/cn/indicator.json?symbol={}&type=all&is_detail=true&count={}&timestamp={}'.format(code,count,timestamp)
    logger.info("Request ur is "+url)
    return fetch(url)

def fetch_cash_flow(code):
    timestamp = round(time.time()*1000)
    url = 'https://stock.xueqiu.com/v5/stock/finance/cn/cash_flow.json?symbol={}&type=all&is_detail=true&count={}&timestamp={}'.format(code,count,timestamp)
    logger.info("Request ur is "+url)
    return fetch(url)

def fetch_quoto(code):
    timestamp = round(time.time()*1000)
    url = 'https://stock.xueqiu.com/v5/stock/quote.json?extend=detail&symbol={}'.format(code)
    logger.info("Request ur is "+url)
    return fetch(url)


# In[8]:


def n(str):
    if str is None:
        return 0
    return str
def percent(obj):
    if obj is None:
        return "-"
    return round(obj,5)
def percent_100(obj):
    if obj is None:
        return "-"
    return str(round(obj*100,2))+"%"
def rnd(obj):
    if obj is None:
        return "-"
    return round(obj,3)
def min8(obj):
    if str is None:
        return  "-"
    return round(obj/10**8,2)


# In[9]:


rst=fetch_quoto(stock_code)
stock_name = rst['data']['quote']['name']
current_price = rst['data']['quote']['current']
price_percent = rst['data']['quote']['percent']


# In[10]:


logger.setLevel(logging.INFO)
rst = fetch_balance(stock_code)
balance_list=rst['data']['list']
report_dict={}
for balance_data in balance_list:
    
    report_date=balance_data['report_date'] #报告Timestamp
    logger.debug("Timestamp"+str(report_date))
    
    report_name=balance_data['report_name'] #报告期
    logger.warning("资产负债表，报告期"+str(report_name))
    
    total_current_assets=n(balance_data['total_current_assets'][0]) #流动资产
    logger.debug("流动资产"+str(total_current_assets))
    total_current_liab=n(balance_data['total_current_liab'][0]) #流动负债
    logger.debug("流动负债"+str(total_current_liab))
    total_holders_equity=n(balance_data['total_holders_equity'][0]) #股东权益
    logger.debug("股东权益"+str(total_holders_equity))

    currency_funds=n(balance_data['currency_funds'][0]) #货币资金
    logger.debug("货币资金"+str(currency_funds))
    tradable_fnncl_assets=n(balance_data['tradable_fnncl_assets'][0]) #交易性金融资产
    logger.debug("交易性金融资产"+str(tradable_fnncl_assets))
    ar_and_br=n(balance_data['ar_and_br'][0]) #应收票据及应收账款
    logger.debug("应收票据及应收账款"+str(ar_and_br))
    pre_payment=n(balance_data['pre_payment'][0]) #预付款项
    logger.debug("预付款项"+str(pre_payment))
    interest_receivable=n(balance_data['interest_receivable'][0]) #应收利息
    logger.debug("应收利息"+str(interest_receivable))
    dividend_receivable=n(balance_data['dividend_receivable'][0]) #应收股利
    logger.debug("应收股利"+str(dividend_receivable))
    othr_receivables=n(balance_data['othr_receivables'][0]) #其他应收款
    logger.info("其他应收款"+str(othr_receivables))
    inventory=n(balance_data['inventory'][0]) #存货
    logger.debug("存货"+str(inventory))
    contractual_assets=n(balance_data['contractual_assets'][0]) #合同资产
    logger.debug("合同资产"+str(contractual_assets))
    to_sale_asset=n(balance_data['to_sale_asset'][0]) #划分为持有待售的资产
    logger.debug("划分为持有待售的资产"+str(to_sale_asset))
    nca_due_within_one_year=n(balance_data['nca_due_within_one_year'][0]) #一年内到期的非流动资产
    logger.debug("一年内到期的非流动资产"+str(nca_due_within_one_year))
    othr_current_assets=n(balance_data['othr_current_assets'][0]) #其他流动资产
    logger.debug("其他流动资产"+str(othr_current_assets))
    
    
    salable_financial_assets=n(balance_data['salable_financial_assets'][0]) #可供出售金融资产
    logger.debug("可供出售金融资产"+str(salable_financial_assets))
    held_to_maturity_invest=n(balance_data['held_to_maturity_invest'][0]) #持有至到期投资
    logger.debug("持有至到期投资"+str(held_to_maturity_invest))
    lt_receivable=n(balance_data['lt_receivable'][0]) #长期应收款
    logger.debug("长期应收款"+str(lt_receivable))
    lt_equity_invest=n(balance_data['lt_equity_invest'][0]) #长期股权投资
    logger.debug("长期股权投资"+str(lt_equity_invest))
    
    other_eq_ins_invest=n(balance_data['other_eq_ins_invest'][0]) #其他权益工具投资
    logger.debug("其他权益工具投资"+str(other_eq_ins_invest))
    other_illiquid_fnncl_assets=n(balance_data['other_illiquid_fnncl_assets'][0]) #其他非流动金融资产
    logger.debug("其他非流动金融资产"+str(other_illiquid_fnncl_assets))
    invest_property=n(balance_data['invest_property'][0]) #投资性房地产
    logger.debug("投资性房地产"+str(invest_property))
    
    fixed_asset_sum= n(balance_data['fixed_asset_sum'][0]) #固定资产合计
    logger.debug("固定资产合计"+str(fixed_asset_sum))
    fixed_assets_disposal= n(balance_data['fixed_assets_disposal'][0]) #固定资产清理
    logger.debug("固定资产清理"+str(fixed_assets_disposal))
    construction_in_process_sum= n(balance_data['construction_in_process_sum'][0]) #在建工程合计
    logger.debug("在建工程合计"+str(construction_in_process_sum))
    project_goods_and_material= n(balance_data['project_goods_and_material'][0]) #工程物资
    logger.debug("工程物资"+str(project_goods_and_material))
    productive_biological_assets= n(balance_data['productive_biological_assets'][0]) #生产性生物资产
    logger.debug("生产性生物资产"+str(productive_biological_assets))
    oil_and_gas_asset= n(balance_data['oil_and_gas_asset'][0]) #油气资产
    logger.debug("油气资产"+str(oil_and_gas_asset))
    intangible_assets= n(balance_data['intangible_assets'][0]) #无形资产
    logger.debug("无形资产"+str(intangible_assets))
    dev_expenditure= n(balance_data['dev_expenditure'][0]) #开发支出
    logger.debug("开发支出"+str(dev_expenditure))
    goodwill= n(balance_data['goodwill'][0]) #商誉
    logger.debug("商誉"+str(goodwill))
    lt_deferred_expense= n(balance_data['lt_deferred_expense'][0]) #长期待摊费用
    logger.debug("长期待摊费用"+str(lt_deferred_expense))
    dt_assets= n(balance_data['dt_assets'][0]) #递延所得税资产
    logger.debug("递延所得税资产"+str(dt_assets))
    othr_noncurrent_assets= n(balance_data['othr_noncurrent_assets'][0]) #其他非流动资产
    logger.debug("其他非流动资产"+str(othr_noncurrent_assets))
    
    total_assets= n(balance_data['total_assets'][0]) #资产合计
    logger.info("资产合计"+str(total_assets))
    
    finance_asset= (tradable_fnncl_assets+interest_receivable+dividend_receivable+to_sale_asset
    +nca_due_within_one_year +salable_financial_assets +held_to_maturity_invest 
    +other_eq_ins_invest+other_illiquid_fnncl_assets +invest_property)#金融资产
    logger.info("金融资产"+str(finance_asset))
    finance_asset_percent=finance_asset/total_assets
    logger.info("金融资产占比"+str(finance_asset_percent))
    
#     business_asset=  (currency_funds+ar_and_br+pre_payment+inventory+contractual_assets+othr_current_assets
#     +lt_receivable+fixed_asset_sum+construction_in_process_sum
#     +project_goods_and_material+productive_biological_assets
#     +productive_biological_assets +oil_and_gas_asset+intangible_assets
#     +dev_expenditure+goodwill+lt_deferred_expense+dt_assets
#     +othr_noncurrent_assets+othr_receivables+fixed_assets_disposal) #经营资产
    
    business_asset = total_assets - finance_asset-lt_equity_invest
    logger.info("经营资产"+str(business_asset))
    business_asset_percent=business_asset/total_assets
    logger.info("经营资产占比"+str(business_asset_percent))
    
    logger.info("长期股权投资"+str(lt_equity_invest))
    lt_equity_invest_percent=lt_equity_invest/total_assets
    logger.info("长期股权占比"+str(lt_equity_invest_percent))
    
    z_business_asset_percent=(total_current_assets-total_current_liab)/total_holders_equity #周期性经营投入占比
    logger.info("周期性经营投入占比"+str(z_business_asset_percent))
    
    long_business_asset=(fixed_asset_sum+construction_in_process_sum
                                 +project_goods_and_material+productive_biological_assets
                                 +productive_biological_assets +oil_and_gas_asset+intangible_assets
                                 +dev_expenditure+goodwill+lt_deferred_expense+dt_assets
                                 +othr_noncurrent_assets+fixed_assets_disposal) #长期经营资产
    logger.info("长期经营资产"+str(long_business_asset))
    
    long_business_asset_percent=long_business_asset/total_assets #长期经营资产占比
    logger.info("长期经营资产占比"+str(long_business_asset_percent))
    
    finance_turnover=total_assets/total_holders_equity #财务杠杆倍数
    logger.info("财务杠杆倍数"+str(finance_turnover))
    
    report_dict[report_date]={} #仅声明一次
    
    report_dict[report_date]["finance_asset"]=finance_asset
    report_dict[report_date]["finance_asset_percent"]=finance_asset_percent
    report_dict[report_date]["business_asset"]=business_asset
    report_dict[report_date]["business_asset_percent"]=business_asset_percent
    report_dict[report_date]["lt_equity_invest"]=lt_equity_invest #长期股权投资
    report_dict[report_date]["lt_equity_invest_percent"]=lt_equity_invest_percent#长期股权投资占比
    report_dict[report_date]["goodwill"]=goodwill #商誉
        
    
    logger.info("资产合计"+str(finance_asset+business_asset+lt_equity_invest))   
    report_dict[report_date]["report_name"]=report_name
    report_dict[report_date]["z_business_asset_percent"]=z_business_asset_percent
    report_dict[report_date]["long_business_asset"]=long_business_asset
    report_dict[report_date]["long_business_asset_percent"]=long_business_asset_percent
    report_dict[report_date]["finance_turnover"]=finance_turnover #财务杠杆倍数
    
    report_dict[report_date]["total_current_assets"]=total_current_assets #流动资产
    report_dict[report_date]["total_current_liab"]=total_current_liab #流动负债
    
logger.debug("report_dict"+str(report_dict))   


# In[11]:


rst = fetch_income(stock_code)
income_list=rst['data']['list']
income_dict={}
for income_data in income_list:
    report_date=income_data['report_date'] #报告Timestamp
    logger.debug("Timestamp"+str(report_date))
    
    report_name=income_data['report_name'] #报告期
    logger.warning("利润表，报告期"+str(report_name))
    day = None
    if '一季报' in report_name:
        day=90
    elif '中报' in report_name:
        day= 181
    elif '三季报' in report_name:
        day= 273
    else:
        day=365
        
    total_revenue=n(income_data['total_revenue'][0]) #营业总收入
    logger.info("营业总收入"+str(total_revenue))
    
    operating_costs=n(income_data['operating_costs'][0]) #营业总成本
    logger.debug("营业总成本"+str(operating_costs))
    
    operating_cost=n(income_data['operating_cost'][0]) #营业成本
    logger.debug("营业成本"+str(operating_cost))
    
    net_profit=n(income_data['net_profit'][0]) #净利润
    logger.info("净利润"+str(net_profit))
    
    
    income_tax_expenses=n(income_data['income_tax_expenses'][0]) #所得税费用
    logger.info("所得税费用"+str(income_tax_expenses))
    
    financing_expenses=n(income_data['financing_expenses'][0]) #财务费用
    logger.info("财务费用"+str(financing_expenses))

    profit_total_amt=n(income_data['profit_total_amt'][0]) #利润总额
    logger.info("利润总额"+str(profit_total_amt))
    net_profit_after_nrgal_atsolc=n(income_data['net_profit_after_nrgal_atsolc'][0]) #扣除非经常性损益后的净利润
    logger.info("扣除非经常性损益后的净利润"+str(net_profit_after_nrgal_atsolc))
  
    business_lever=(total_revenue-operating_cost)/(total_revenue-operating_costs) #经营杠杆系数
    logger.debug("经营杠杆系数"+str(business_lever))
    
    per_benefit_percent=(net_profit+income_tax_expenses+financing_expenses)/(total_revenue) #息税前经营利润率
    logger.debug("息税前经营利润率"+str(per_benefit_percent))
    
    times= (365/day)
    logger.debug("倍率"+str(times))
        
    long_asset_turnover_percent=(total_revenue)/(report_dict[report_date]["long_business_asset"])*times #长期资产周转率
    logger.debug("长期资产周转率"+str(long_asset_turnover_percent))
    
    business_turnover_percent=(total_revenue)/(report_dict[report_date]["total_current_assets"]
                                               -report_dict[report_date]["total_current_liab"]) *times#周转性经营投入周转率
    logger.debug("周转性经营投入周转率"+str(business_turnover_percent))
    
    financial_cost_percent=(net_profit+income_tax_expenses)/(net_profit+income_tax_expenses+financing_expenses) #财务成本效应比率
    logger.info("财务成本效应比率"+str(financial_cost_percent))
    
    tax_cost_percent= net_profit/(net_profit+income_tax_expenses)#企业所得税效应比率
    logger.info("企业所得税效应比率"+str(tax_cost_percent))
    
    real_benefit_percent = net_profit_after_nrgal_atsolc/profit_total_amt # 扣非净利率占比
    logger.info("扣非净利率占比"+str(real_benefit_percent))
    
    report_dict[report_date]["total_current_assets"]=total_current_assets #流动资产
    report_dict[report_date]["total_current_liab"]=total_current_liab #流动负债
    
    report_dict[report_date]["total_revenue"]=total_revenue #营业总收入
    report_dict[report_date]["net_profit"]=net_profit #净利润
    report_dict[report_date]["profit_total_amt"]=profit_total_amt #利润总额
    report_dict[report_date]["business_lever"]=business_lever
    report_dict[report_date]["per_benefit_percent"]=per_benefit_percent
    report_dict[report_date]["long_asset_turnover_percent"]=long_asset_turnover_percent
    report_dict[report_date]["business_turnover_percent"]=business_turnover_percent
    report_dict[report_date]["financial_cost_percent"]=financial_cost_percent #财务成本效应比率
    report_dict[report_date]["tax_cost_percent"]=tax_cost_percent #企业所得税效应比率
    report_dict[report_date]["real_benefit_percent"]=real_benefit_percent#扣非净利率占比

# In[12]:

rst = fetch_indicator(stock_code)
indicator_list=rst['data']['list']
indicator_dict={}
for indicator_data in indicator_list:
    report_date=indicator_data['report_date'] #报告Timestamp
    logger.debug("Timestamp"+str(report_date))
    
    report_name=indicator_data['report_name'] #报告期
    logger.warning("主要指标，报告期"+str(report_name))
    day = None
    
    if '一季报' in report_name:
        day=90
    elif '中报' in report_name:
        day= 181
    elif '三季报' in report_name:
        day= 273
    else:
        day=365
    
    logger.debug("日期"+str(day))
    
    cash_cycle=indicator_data['cash_cycle'][0] #现金循环周期
    logger.debug("现金循环周期"+str(cash_cycle))
    
    fixed_asset_turnover_ratio=day/n(indicator_data['fixed_asset_turnover_ratio'][0]) #固定资产周转天数
    logger.debug("固定资产周转天数"+str(fixed_asset_turnover_ratio))
    
    report_dict[report_date]["cash_cycle"]= cash_cycle #现金周期
    logger.info("现金周期"+str(report_dict[report_date]["cash_cycle"]))
    
    report_dict[report_date]["fixed_asset_turnover_ratio"]=fixed_asset_turnover_ratio
    logger.info("固定资产周转天数"+str(fixed_asset_turnover_ratio))


# In[13]:


rst = fetch_cash_flow(stock_code)
cash_flow_list=rst['data']['list']
cash_flow_dict={}
for cash_flow_data in cash_flow_list:
    report_date=cash_flow_data['report_date'] #报告Timestamp
    logger.debug("Timestamp"+str(report_date))
    
    report_name=cash_flow_data['report_name'] #报告期
    logger.warning("现金流表，报告期"+str(report_name))
    
    cash_received_of_sales_service=n(cash_flow_data['cash_received_of_sales_service'][0]) #销售商品、提供劳务收到的现金
    logger.debug("销售商品、提供劳务收到的现金"+str(cash_received_of_sales_service))
    
    ncf_from_oa=n(cash_flow_data['ncf_from_oa'][0]) #经营活动产生的现金流量净额
    logger.info("经营活动产生的现金流量净额"+str(ncf_from_oa))
    
    cash_paid_for_assets=n(cash_flow_data['cash_paid_for_assets'][0]) #购建固定资产、无形资产和其他长期资产支付的现金
    logger.debug("购建固定资产、无形资产和其他长期资产支付的现金"+str(cash_paid_for_assets))
    
    net_cash_of_disposal_assets=n(cash_flow_data['net_cash_of_disposal_assets'][0]) #处置固定资产、无形资产和其他长期资产收回的现金净额
    logger.debug("处置固定资产、无形资产和其他长期资产收回的现金净额"+str(net_cash_of_disposal_assets))

    net_cash_amt_from_branch=n(cash_flow_data['net_cash_amt_from_branch'][0]) #取得子公司及其他营业单位支付的现金净额
    logger.debug("取得子公司及其他营业单位支付的现金净额"+str(net_cash_amt_from_branch))

    net_cash_of_disposal_branch=n(cash_flow_data['net_cash_of_disposal_branch'][0]) #处置子公司及其他营业单位收到的现金净额
    logger.debug("处置子公司及其他营业单位收到的现金净额"+str(net_cash_of_disposal_branch))
    
    
    benefit_percentage= ncf_from_oa/report_dict[report_date]["profit_total_amt"] #净利润检验含金量
    logger.debug("净利润检验含金量"+str(benefit_percentage))
    op_percentage= cash_received_of_sales_service/report_dict[report_date]["total_revenue"]#营业收入含金量检验
    logger.debug("营业收入含金量检验"+str(op_percentage))
    
    net_investment= cash_paid_for_assets-net_cash_of_disposal_assets#长期经营资产净投资额
    logger.info("长期经营资产净投资额"+str(net_investment))
    
    # 这里折旧嫌麻烦就默认注释掉了，也可以打开，每次输入年报中的折旧金额
    depreciation = 0
    # if "年报" in report_name:
    #     depreciation = input('请输入 {} 折旧.\n'.format(report_name))
    #     if len(depreciation) == 0:
    #         depreciation=0
    #     print(report_name+"折旧金额： "+str(depreciation))
    # else:
    #     depreciation = 0
    
    out_lay= net_investment-float(depreciation) #长期经营资产扩张性资本支出
    logger.debug("长期经营资产扩张性资本支出"+str(out_lay))
    out_lay_percent=  out_lay/  report_dict[report_date]["long_business_asset"] #长期经营资产扩张性资本支出比例
    logger.debug("长期经营资产扩张性资本支出比例"+str(out_lay_percent))

    net_merger=  net_cash_amt_from_branch - net_cash_of_disposal_branch #并购活动净合并额
    logger.debug("并购活动净合并额"+str(net_merger))
    
    cash_satify_percent=  ncf_from_oa /  (net_investment+net_merger) #现金自给率
    logger.debug("现金自给率"+str(cash_satify_percent))
    
    report_dict[report_date]["benefit_percentage"]=benefit_percentage #净利润检验含金量
    report_dict[report_date]["op_percentage"]=op_percentage #营业收入含金量检验
    report_dict[report_date]["net_investment"]=net_investment #长期经营资产净投资额
    report_dict[report_date]["out_lay"]=out_lay #长期经营资产扩张性资本支出
    report_dict[report_date]["out_lay_percent"]=out_lay_percent #长期经营资产扩张性资本支出比例
    report_dict[report_date]["net_merger"]=net_merger #并购活动净合并额
    report_dict[report_date]["cash_satify_percent"]=cash_satify_percent #现金自给率


# In[14]:

wb = openpyxl.load_workbook('财务数据模板.xlsx')
sht=wb['公司分析']

green_font = Font(size=16,color="00a600",bold=True)
red_font = Font(size=16,color="ff0000",bold=True)

del_font = Font(strike=True)

if price_percent>0:
    sht.cell(1,2).font=red_font
else:
    sht.cell(1,2).font=green_font

sht['B1']='{}  {} ( {}% )'.format(stock_name, current_price, price_percent)

###
i = 0
for report_data in report_dict.values():
    sht.cell(1,i+3).value = report_data['report_name']
    sht.cell(2,i+3).value = percent(report_data['z_business_asset_percent'])
    sht.cell(3,i+3).value = percent(report_data['long_business_asset_percent'])
    sht.cell(4,i+3).value = rnd(report_data['fixed_asset_turnover_ratio'])    
    if report_data['long_business_asset_percent']>0.5:
         sht.cell(5,i+3).value='重资产'
    else:
        sht.cell(5,i+3).value='轻资产'
        sht.cell(4,i+3).font=del_font
        sht.cell(14,i+3).font=del_font
    sht.cell(6,i+3).value = str(min8(report_data['finance_asset']))+" ("+percent_100(report_data['finance_asset_percent'])+")"
    sht.cell(7,i+3).value =  str(min8(report_data['lt_equity_invest']))+" ("+percent_100(report_data['lt_equity_invest_percent'])+")"
    sht.cell(8,i+3).value =  str(min8(report_data['business_asset']))+" ("+percent_100(report_data['business_asset_percent'])+")"
    sht.cell(10,i+3).value = min8(report_data['goodwill'])
    sht.cell(11,i+3).value = rnd(report_data['business_lever'])
    sht.cell(12,i+3).value = rnd(report_data['cash_cycle'])
    sht.cell(13,i+3).value = percent(report_data['per_benefit_percent'])
    
    sht.cell(14,i+3).value = percent( report_data["long_asset_turnover_percent"])
    sht.cell(15,i+3).value = percent(report_data["business_turnover_percent"])
    
    sht.cell(16,i+3).value = percent( report_data["financial_cost_percent"])
    sht.cell(17,i+3).value = rnd( report_data["finance_turnover"])
    sht.cell(18,i+3).value = percent( report_data["tax_cost_percent"])
    sht.cell(19,i+3).value = rnd( report_data["benefit_percentage"])
    sht.cell(20,i+3).value = rnd( report_data["op_percentage"])
    sht.cell(21,i+3).value = percent( report_data["real_benefit_percent"])
    
    sht.cell(22,i+3).value = min8(report_data["net_investment"])
    sht.cell(23,i+3).value = min8(report_data["out_lay"])
    sht.cell(24,i+3).value = percent(report_data["out_lay_percent"])

    sht.cell(25,i+3).value = min8(report_data["net_merger"])
    
    sht.cell(26,i+3).value = percent(report_data["cash_satify_percent"])
    
    sht.cell(29,i+3).value = min8(report_data["long_business_asset"])
    
    sht.cell(2,i+3).number_format='0.00%'
    sht.cell(3,i+3).number_format='0.00%'
    sht.cell(13,i+3).number_format='0.00%'
    sht.cell(14,i+3).number_format='0.00%'
    sht.cell(15,i+3).number_format='0.00%'
    sht.cell(16,i+3).number_format='0.00%'
    sht.cell(18,i+3).number_format='0.00%'
    sht.cell(21,i+3).number_format='0.00%'
    sht.cell(24,i+3).number_format='0.00%'
    sht.cell(26,i+3).number_format='0.00%'
    sht.cell(1,i+3).alignment = Alignment(horizontal='center', vertical='center')
    sht.cell(5,i+3).alignment = Alignment(horizontal='right', vertical='center')
    sht.cell(6,i+3).alignment = Alignment(horizontal='right', vertical='center')
    sht.cell(7,i+3).alignment = Alignment(horizontal='right', vertical='center')
    sht.cell(8,i+3).alignment = Alignment(horizontal='right', vertical='center')
    i+=1
###
stock_name=stock_name.replace("*", "")

#这里的保存位置可以修改
wb.save('./'+stock_name+'-财务报表.xlsx')




